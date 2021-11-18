from openpyxl import load_workbook
import numpy as np
import os.path
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.font_manager

class bcolors:
    """
    Reference from: 
    https://svn.blender.org/svnroot/bf-blender/trunk/blender/build_files/scons/tools/bcolors.py
    """
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def get_lists(tab=None, cfg="ubrain"):
    """
    List format = [BUF, RNG, CNT, CMP, PC, REST, TOTAL]
    Units:
        - Area: um^2
        - Power: uW
        - Runtime: ms
        - Energy: nJ
    Returns: 
        area_list, dynamic_list, leakage_list, power_list, runtime, energy_list
    """
    runtime = 14.0/128*10**(3)
    if cfg == 'ubrain':
        area = get_cell_vals(tab, 'G29:G35')
        dynamic = get_cell_vals(tab, 'H29:H35')
        leakage = get_cell_vals(tab, 'I29:I35')
        power = get_cell_vals(tab, 'J29:J35')
        energy = (np.asarray(runtime) * power).tolist()
        return area, dynamic, leakage, power, runtime, energy
    elif cfg == "sc": 
        area = get_cell_vals(tab, 'P29:P35')
        dynamic = get_cell_vals(tab, 'Q29:Q35')
        leakage = get_cell_vals(tab, 'R29:R35')
        power = get_cell_vals(tab, 'S29:S35')
        energy = (np.asarray(runtime) * power).tolist()
        return area, dynamic, leakage, power, runtime, energy
    else:
        print(bcolors.FAIL + f'Unrecognized cfg {cfg}. Options: sc, ubrain' + bcolors.ENDC)
        exit()
    
def get_cell_vals(tab, str_range):
    if ':' in str_range:
        cells = tab[str_range]
        return [x.value for cell in cells for x in cell]
    else:
        cell = tab[str_range]
        return cell.value

def query(str_tab=None, cfg='ubrain', absolute_path=None, filename=None):
    wb = load_workbook(filename=absolute_path+filename, data_only=True)
    sheets = wb.sheetnames
    ind = sheets.index(str_tab)
    return get_lists(tab=wb[sheets[ind]], cfg=cfg)

def query_workbook(workbook=None, str_tab=None, cfg='ubrain'):
    sheets = workbook.sheetnames
    ind = sheets.index(str_tab)
    return get_lists(tab=workbook[sheets[ind]], cfg=cfg)

def get_sheetname(absolute_path = None, filename = None):
    wb = load_workbook(filename=absolute_path+filename, data_only=True)
    return wb.sheetnames

def get_sheetname_workbook(workbook = None):
    return workbook.sheetnames

def power_area_fold_stacked_bar(conv_num=1, design='ubrain', absolute_path=None, filename=None):
    font = {'family':'Times New Roman', 'size': 6}
    matplotlib.rc('font', **font)

    lkg = []; dym = []; area = []

    cfg = 'conv'+str(conv_num)
    workbook = load_workbook(filename=absolute_path+filename, data_only=True)
    for i in get_sheetname_workbook(workbook):
        if cfg in i:
            area_conv, dynamic_conv, leakage_conv, _, _, _ = query_workbook(workbook, i, design)
            dym.append(dynamic_conv[len(dynamic_conv)-1])
            lkg.append(leakage_conv[len(leakage_conv)-1])
            area.append(area_conv[len(area_conv)-1])
        else: pass

    # Sanity check
    if conv_num == 1: 
        assert len(lkg) == 5
        x_axis = ['H0', 'H1', 'H2', 'H3', 'H4']
    elif conv_num == 2: 
        assert len(lkg) == 6
        x_axis = ['H0', 'H1', 'H2', 'H3', 'H4', 'H5']
    else:
        print(bcolors.FAIL + f'Unrecognized conv layer number {conv_num}.' + bcolors.FAIL)
        exit()
    
    # plot double axis
    # plot area and power
    my_dpi = 300
    if conv_num == 1:
        fig_h = 1.1
    else:
        fig_h = 1
    fig_w = 3.3115

    # TODO: Change color
    area_color = "#AAAAAA"
    dyn_color = "#FF7F7F"
    lkg_color = "#7A81FF"

    x_idx = np.arange(len(x_axis))

    width = 0.3

    fig, area_ax = plt.subplots(figsize=(fig_w, fig_h))
    area_ax.bar(x_idx - 0.5 * width, area, width, hatch = None, alpha=0.99, color=area_color, label='Area')
    area_ax.set_ylabel('Area ($mm^2$)')
    area_ax.minorticks_off()

    power_ax = area_ax.twinx()
    power_ax.bar(x_idx + 0.5 * width, dym, width, hatch = None, alpha=0.99, color=dyn_color, label='Dynamic')
    power_ax.bar(x_idx + 0.5 * width, lkg, width, bottom=dym, hatch = None, alpha=0.99, color=lkg_color, label='Leakage')

    if conv_num == 1:
        bars, labels = area_ax.get_legend_handles_labels()
        bars2, labels2 = power_ax.get_legend_handles_labels()
        if design == 'ubrain':
            area_ax.legend(bars + bars2, labels + labels2, loc="upper center", ncol=3, frameon=True)

    power_ax.set_ylabel('Power ($mW$)')
    power_ax.minorticks_off()

    area_ax.set_xticks(x_idx)
    area_ax.set_xticklabels(x_axis)
    plt.xlim(x_idx[0]-0.5, x_idx[-1]+0.5)
    plt.yscale("linear")
    
    print("area_ax ylim: ", area_ax.get_ylim())
    print("power_ax ylim: ", power_ax.get_ylim())

    # Manually adjust axis limit & ticks
    if design == 'ubrain':
        if conv_num == 2:
            area_ax.set_ylim((0, 24))
            area_ax.set_yticks((0, 10, 20))
            area_ax.set_yticklabels(("{:2d}".format(0), "{:2d}".format(10), "{:2d}".format(20)))
            power_ax.set_ylim((0, 24))
            power_ax.set_yticks((0, 10, 20))
            power_ax.set_yticklabels(("{:2d}".format(0), "{:2d}".format(10), "{:2d}".format(20)))
        else:
            area_ax.set_ylim((0, 2.3))
            area_ax.set_yticks((0, 0.5, 1.0, 1.5))
            area_ax.set_yticklabels(("{:2d}".format(0), "{:.1f}".format(0.5), "{:.1f}".format(1.0), "{:.1f}".format(1.5)))
            power_ax.set_ylim((0, 2.3))
            power_ax.set_yticks((0, 0.5, 1.0, 1.5))
            power_ax.set_yticklabels(("{:2d}".format(0), "{:.1f}".format(0.5), "{:.1f}".format(1.0), "{:.1f}".format(1.5)))
    else: #sc
        if conv_num == 1:
            area_ax.set_ylim((0, 5))
            area_ax.set_yticks((0, 2, 4))
            area_ax.set_yticklabels(("{:2d}".format(0), "{:2d}".format(2), "{:2d}".format(4)))
            power_ax.set_ylim((0, 5))
            power_ax.set_yticks((0, 2, 4))
            power_ax.set_yticklabels(("{:2d}".format(0), "{:2d}".format(2), "{:2d}".format(4)))
        else:
            area_ax.set_ylim((0, 130))
            area_ax.set_yticks((0, 50, 100))
            area_ax.set_yticklabels(("{:2d}".format(0), "{:2d}".format(50), "{:2d}".format(100)))
            power_ax.set_ylim((0, 105))
            power_ax.set_yticks((0, 50, 100))
            power_ax.set_yticklabels(("{:2d}".format(0), "{:2d}".format(50), "{:2d}".format(100)))

    fig.tight_layout()
    plt.savefig(absolute_path + f'conv{conv_num}_HWreuse_{design}.pdf', bbox_inches='tight', dpi=my_dpi, pad_inches=0.02)


def main():
    # NOTE: Update absolute path and filename
    path = '/home/zhewen/Repo/UnarySim/app/uBrain/hw/data_extract/'
    filename = "uBrain_resource.xlsx"
    file_exists = os.path.exists(path+filename)
    if file_exists == False:
        print(bcolors.FAIL + f'{filename} does not exist at path {path}. Did you forget to put it in?' + bcolors.ENDC)
        exit()
    else: print(bcolors.OKGREEN + f'Processing {path+filename}...' + bcolors.ENDC)
    
    power_area_fold_stacked_bar(1, 'ubrain', path, filename)
    power_area_fold_stacked_bar(2, 'ubrain', path, filename)

    power_area_fold_stacked_bar(1, 'sc', path, filename)
    power_area_fold_stacked_bar(2, 'sc', path, filename)

    # example query_workbook
    # workbook = load_workbook(filename=path+filename, data_only=True)
    # print(query_workbook(workbook, 'conv1-F1', 'ubrain'))
    # print(query_workbook(workbook, 'conv1-F1', 'sc'))
    # print(query_workbook(workbook, 'conv1-F2', 'ubrain'))
    # print(query_workbook(workbook, 'conv1-F2', 'sc'))
    # print(query_workbook(workbook, 'conv1-F4', 'ubrain'))
    # print(query_workbook(workbook, 'conv1-F4', 'sc'))
    # print(query_workbook(workbook, 'conv1-F8', 'ubrain'))
    # print(query_workbook(workbook, 'conv1-F8', 'sc'))
    # print(query_workbook(workbook, 'conv1-F16', 'ubrain'))
    # print(query_workbook(workbook, 'conv1-F16', 'sc'))

    # print(query_workbook(workbook, 'conv2-F1', 'ubrain'))
    # print(query_workbook(workbook, 'conv2-F1', 'sc'))
    # print(query_workbook(workbook, 'conv2-F2', 'ubrain'))
    # print(query_workbook(workbook, 'conv2-F2', 'sc'))
    # print(query_workbook(workbook, 'conv2-F4', 'ubrain'))
    # print(query_workbook(workbook, 'conv2-F4', 'sc'))
    # print(query_workbook(workbook, 'conv2-F8', 'ubrain'))
    # print(query_workbook(workbook, 'conv2-F8', 'sc'))
    # print(query_workbook(workbook, 'conv2-F16', 'ubrain'))
    # print(query_workbook(workbook, 'conv2-F16', 'sc'))
    # print(query_workbook(workbook, 'conv2-F32', 'ubrain'))
    # print(query_workbook(workbook, 'conv2-F32', 'sc'))

    # print(query_workbook(workbook, 'fc3-F256', 'ubrain'))
    # print(query_workbook(workbook, 'fc3-F256', 'sc'))

    # print(query_workbook(workbook, 'rnn4-F1', 'ubrain'))
    # print(query_workbook(workbook, 'rnn4-F1', 'sc'))

    # print(query_workbook(workbook, 'fc5-F1', 'ubrain'))
    # print(query_workbook(workbook, 'fc5-F1', 'sc'))
    
    # print(query_workbook(workbook, 'fc6-F1', 'ubrain'))
    # print(query_workbook(workbook, 'fc6-F1', 'sc'))


if __name__ == "__main__":
    main()
